#!/usr/bin/env python3
"""
Build data.json for the biopharma capital-flow dashboard.

Reads the manually-curated trackers, cleans the messy human-entered
values (currency strings, disclosure placeholders, dates, investor
columns spread across many fields) and writes a single data.json the
static dashboard consumes.

SOURCE OF TRUTH
---------------
USE_SHEETS = True pulls each tab from Google Sheets "published to the web"
as CSV -- no API key, no billing, no card. Each tab's CSV lives at:
  https://docs.google.com/spreadsheets/d/e/<PUBKEY>/pub?gid=<GID>&single=true&output=csv
The <PUBKEY> is per-document (already filled below from your publish links).
The <GID> is per-tab -- fill the GID map below (one number per tab).
  How to get a gid: open the Sheet, click a tab; the address bar shows
  ".../edit#gid=123456789" -- that number is the gid for that tab.

USE_SHEETS = False falls back to the local .xlsx files (for testing).
Parsing is identical either way.
"""
import json, math, re, datetime
import pandas as pd

# --- config -----------------------------------------------------------
USE_SHEETS = True            # True = pull from Google Sheets; False = local xlsx
EUR_USD = 1.08               # fixed FX for aggregate $ views (edit as needed)
GBP_USD = 1.27

# "Publish to web" document keys (from your two links)
PUBKEY = {
    "deal":      "2PACX-1vRmcRt9FRyllvcfIe1EkEj9sok91UjAJPVkAFRvC9gr2H191WlnV5SWljkU7AFrritTMt1zcJ3BttYq",
    "financing": "2PACX-1vQDmKjET-kgICjQ2I0ZHXuq5LvacHUcnVjELu_pAbOar1lPMVMknwlDUJR30s17TioXVxvuRXQH0feX",
}

# >>> gids per tab (from your Sheets) <<<
GID = {
    "Deals":           131319432,    # Deal tracker doc
    "Licensing":       0,            # Deal tracker doc
    "Fundraises":      0,            # Financing doc
    "IPO":             668188943,    # Financing doc
    "Reverse mergers": 1349784264,   # Financing doc
    "Royalty deals":   1049501359,   # Financing doc
}

# which document each tab lives in, + local-file fallback (testing only)
DOC   = {"Deals": "deal", "Licensing": "deal",
         "Fundraises": "financing", "IPO": "financing",
         "Reverse mergers": "financing", "Royalty deals": "financing"}
LOCAL = {"Deals": ("Deal_tracker_062426.xlsx", "M&A"),
         "Licensing": ("Deal_tracker_062426.xlsx", "Licensing"),
         "Fundraises": ("Fundraise_tracker_062426.xlsx", "Fundraises"),
         "IPO": ("Fundraise_tracker_062426.xlsx", "IPO"),
         "Reverse mergers": ("Fundraise_tracker_062426.xlsx", "Reverse mergers")}
CATEGORIES = list(GID.keys())
# capital "side" for the deals-vs-financings split (auto-scales to new tabs)
SIDE = {c: ("deal" if DOC[c] == "deal" else "raise") for c in CATEGORIES}

def csv_url(cat):
    return (f"https://docs.google.com/spreadsheets/d/e/{PUBKEY[DOC[cat]]}"
            f"/pub?gid={GID[cat]}&single=true&output=csv")

# investor columns differ by sheet; list every column that can hold one
INVESTOR_COLS = ["Lead 1", "Lead 2", "Investor 1", "Investor 2", "Investor 3",
                 "Investor 4", "Investor 5", "Investor 6", "Investor 7", "Investor 8"]
LEAD_COLS = {"Lead 1", "Lead 2", "Investor 1"}   # treated as lead / notable
NULLISH = {"", "n/a", "na", "not disclosed", "nan", "none", "-", "tbd", "undisclosed"}
URL_COLS = {"URL", "Url", "url", "Source URL", "Source url", "Link URL"}

# find the header row by matching known columns -- resilient to a blank
# first row or a shifted header in any tab
KEY_HEADERS = {
    "Deals": {"Buyer", "Seller"}, "Licensing": {"Buyer", "Seller"},
    "Fundraises": {"Company", "Round"}, "IPO": {"Company", "Amount"},
    "Reverse mergers": {"Company", "Target"},
    # "Royalty deals" intentionally omitted -> falls back to first populated row
}

# ----------------------------------------------------------------------
def _tidy(df):
    df = df.dropna(how="all")
    df.columns = [str(c).strip() for c in df.columns]
    df = df.loc[:, ~df.columns.str.startswith("Unnamed")]
    df = df.loc[:, df.columns != ""]
    return df

def load(cat):
    if USE_SHEETS:
        if not GID.get(cat):
            raise SystemExit(f"[build] No gid set for '{cat}'. Fill the GID map in build.py.")
        raw = pd.read_csv(csv_url(cat), header=None, dtype=str)
        if raw.empty or raw.astype(str).apply(
                lambda c: c.str.contains("<html|<!DOCTYPE", case=False, na=False)).any().any():
            raise SystemExit(f"[build] '{cat}' came back as HTML, not CSV -- "
                             f"check its gid and that the tab is published to the web.")
        keys, h = KEY_HEADERS.get(cat, set()), None
        for i in range(min(6, len(raw))):
            vals = {str(x).strip() for x in raw.iloc[i]}
            if keys and len(keys & vals) >= 2:
                h = i; break
        if h is None:   # unknown tab: first row with >=2 non-empty cells is the header
            for i in range(min(6, len(raw))):
                if sum(1 for x in raw.iloc[i] if str(x).strip() and str(x) != "nan") >= 2:
                    h = i; break
            h = h or 0
        cols = ["Unnamed" if (pd.isna(x) or not str(x).strip()) else str(x).strip()
                for x in raw.iloc[h]]
        df = raw.iloc[h + 1:].copy()
        df.columns = cols
        return _tidy(df)
    if cat not in LOCAL:          # e.g. a Sheets-only tab, in local test mode
        return pd.DataFrame()
    f, sheet = LOCAL[cat]
    return _tidy(pd.read_excel(f, sheet_name=sheet, header=1))

def source_urls(cat, id_header):
    """Map identifier -> source article URL from the Source cell hyperlinks.
    (Excel-only; when reading Google Sheets CSV, add a plain 'URL' column
    instead and it flows through load() as a normal field.)"""
    if USE_SHEETS or cat not in LOCAL:
        return {}
    f, sheet = LOCAL[cat]
    from openpyxl import load_workbook
    ws = load_workbook(f).worksheets
    ws = next(w for w in ws if w.title == sheet)
    header = {c.value: c.column for c in ws[2]}           # header on row 2
    idcol = header.get(id_header)
    out = {}
    for row in ws.iter_rows(min_row=3):
        link = next((c.hyperlink.target for c in row if c.hyperlink), None)
        idv = ws.cell(row=row[0].row, column=idcol).value if idcol else None
        if link and idv:
            out[str(idv).strip()] = link
    return out

def clean_str(v):
    if v is None or (isinstance(v, float) and math.isnan(v)):
        return None
    s = str(v).strip()
    return None if s.lower() in NULLISH else s

def parse_money(v):
    """'$1.1B' / '€200M' -> dict with USD-millions magnitude, or None."""
    s = clean_str(v)
    if s is None:
        return None
    cur = "USD"
    if "€" in s or "eur" in s.lower(): cur = "EUR"
    elif "£" in s or "gbp" in s.lower(): cur = "GBP"
    m = re.search(r"([\d,]+(?:\.\d+)?)\s*([bBmMkK])?", s.replace(",", ""))
    if not m:
        return None
    num = float(m.group(1))
    unit = (m.group(2) or "M").upper()
    musd_native = num * {"B": 1000, "M": 1, "K": 0.001}[unit]
    fx = {"USD": 1, "EUR": EUR_USD, "GBP": GBP_USD}[cur]
    return {"raw": s, "currency": cur, "musd": round(musd_native * fx, 2),
            "musd_native": round(musd_native, 2)}

def parse_date(v):
    if isinstance(v, (datetime.datetime, datetime.date, pd.Timestamp)):
        return pd.Timestamp(v).date().isoformat()
    s = clean_str(v)
    if s is None:
        return None
    try:
        return pd.Timestamp(s).date().isoformat()
    except Exception:
        return None

def investors_from(row):
    out = []
    for col in INVESTOR_COLS:
        if col not in row:
            continue
        name = clean_str(row[col])
        if name is None:
            continue
        if name.lower().startswith("other"):   # 'Others (not disclosed)'
            continue
        out.append({"name": name, "lead": col in LEAD_COLS})
    return out

# ----------------------------------------------------------------------
records = {}
investor_rows = []   # long table: one row per (event, investor)

ID_HEADER = {"Deals": "Seller", "Licensing": "Seller", "Fundraises": "Company",
             "IPO": "Company", "Reverse mergers": "Company", "Royalty deals": "Company"}

for cat in CATEGORIES:
    df = load(cat)
    urls = source_urls(cat, ID_HEADER[cat])
    url_col = next((c for c in df.columns if c in URL_COLS), None)
    recs = []
    for _, row in df.iterrows():
        r = {c: clean_str(row[c]) for c in df.columns}
        r["category"] = cat
        r["side"] = SIDE[cat]           # 'deal' or 'raise' — drives the split
        idv = clean_str(row.get(ID_HEADER[cat]))
        # source link: explicit URL column (Sheets) or extracted hyperlink (Excel)
        r["url"] = (clean_str(row[url_col]) if url_col else None) \
                   or (urls.get(idv) if idv else None)
        # dates
        for dcol in ("Date", "Listing date", "IPO date"):
            if dcol in df.columns:
                r[dcol] = parse_date(row[dcol])
        r["date"] = r.get("Date") or r.get("Listing date")
        # money fields (broad net so new tabs like Royalty deals work too)
        MONEY = ("Upfront", "Total Deal Value", "Deal Value", "Amount",
                 "Private placement", "Value", "Total Value")
        for mcol in MONEY:
            if mcol in df.columns:
                r[mcol + "_parsed"] = parse_money(row[mcol])
        # headline value used for size buckets / time series
        headline = next((r[m + "_parsed"] for m in
                         ("Total Deal Value", "Deal Value", "Amount",
                          "Private placement", "Value", "Total Value", "Upfront")
                         if r.get(m + "_parsed")), None)
        r["value_musd"] = headline["musd"] if headline else None
        # investors
        invs = investors_from(row)
        r["investors"] = invs
        counter = r.get("Company") or r.get("Seller") or "?"
        for iv in invs:
            investor_rows.append({"category": cat, "date": r["date"],
                                  "company": counter, "value_musd": r["value_musd"],
                                  **iv})
        recs.append(r)
    records[cat] = recs

data = {
    "generated": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
    "fx": {"EUR_USD": EUR_USD, "GBP_USD": GBP_USD},
    "categories": CATEGORIES,      # tab order for the dashboard
    "side": SIDE,                  # category -> 'deal' | 'raise'
    "records": records,
    "investors": investor_rows,
}
with open("data.json", "w") as fh:
    json.dump(data, fh, indent=2, default=str)

n = sum(len(v) for v in records.values())
print(f"wrote data.json: {n} events across {len(records)} categories, "
      f"{len(investor_rows)} investor rows")
